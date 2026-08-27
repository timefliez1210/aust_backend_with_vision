#!/usr/bin/env python3
"""
Build the Abgleich workbook Alex reviews before the register import.

Nothing is ever written to the database straight from his spreadsheet. This script
compares his Rechnungsausgangsbuch against what production holds and produces ONE
Excel file in which every difference is already decided — he only has to confirm.

    python3 scripts/register-import/build_review_workbook.py \
        --excel "Rechnungsausgangsbuch 2024.xlsx" \
        --sheet 2026 \
        --prod-invoices prod2026.txt \
        --prod-customers prodcust.txt \
        --out Abgleich_2026.xlsx

The two prod dumps are pipe-separated psql output; see README.md for the queries.

Design note — why an .xlsx and not a CSV: Alex works in Excel. Colour tells him at a
glance which of 86 rows need him (about 35), the dropdowns stop him inventing a third
answer we cannot parse, and the frozen header keeps the columns readable while he
scrolls. A CSV would have made him do the triage himself.
"""

import argparse
import datetime as dt
import re
import unicodedata

import openpyxl
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ── Farben ───────────────────────────────────────────────────────────────────
# Deliberately pale: the fill marks a row, it must not drown the text on it.
GREEN = PatternFill("solid", fgColor="E8F5E9")   # passt — nichts zu tun
BLUE = PatternFill("solid", fgColor="E3F2FD")    # fehlt im System — wird angelegt
YELLOW = PatternFill("solid", fgColor="FFF8E1")  # Konflikt — bitte prüfen
RED = PatternFill("solid", fgColor="FFEBEE")     # blockiert — nur Alex kann das
HEAD = PatternFill("solid", fgColor="263238")
THIN = Side(style="thin", color="CFD8DC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

ACTION_FILL = {"OK": GREEN, "NEU": BLUE, "PRÜFEN": YELLOW, "KLÄREN": RED}

# Amounts differing by less than this are the same number written twice — Alex's
# sheet rounds, and a one-cent gap is not a conflict worth his attention.
CENT_TOLERANCE = 100


# ── Normalisierung ───────────────────────────────────────────────────────────

PREFIX = re.compile(
    r"^\s*(anzahlung(\s*\d+\s*%)?|restzahlung|restrechnung|gutschrift|einlagerung|lagerung)\s*",
    re.I,
)
NOISE = re.compile(
    r"\b(gmbh|co|kg|ag|mbh|u|und|ll\.?m\.?|systeme?|dienstleistungen|transport|"
    r"logistik|stiftung|praxis|dr|vertriebs?|deutschland)\b"
)


def strip_prefix(name: str) -> str:
    """`"Anzahlung 30% Kopka"` → `"Kopka"`. Alex encodes the invoice type in the name."""
    return PREFIX.sub("", name).strip()


def kind_of(name: str) -> str:
    """The invoice type Alex encoded in the customer name, if any."""
    m = PREFIX.match(name)
    if not m:
        return "full"
    head = m.group(1).lower()
    if head.startswith("anzahlung"):
        return "partial_first"
    if head.startswith(("restzahlung", "restrechnung")):
        return "partial_final"
    if head.startswith("gutschrift"):
        return "gutschrift"
    return "lagerung"


def tokens(name: str) -> set:
    """Comparable word set: lowercased, accent-stripped, legal-form words removed."""
    s = unicodedata.normalize("NFD", name.lower())
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = NOISE.sub(" ", s)
    return {t for t in re.split(r"[^a-z0-9]+", s) if len(t) > 2}


def eur(cents: int) -> str:
    """German money, for a sentence Alex reads: 6.170,00 €."""
    return f"{cents / 100:,.2f} €".translate(str.maketrans({",": ".", ".": ","}))


def best_customer(name: str, customers):
    """Closest prod customer to `name`, as `(id, name, score)`; score 0 means none."""
    want = tokens(name)
    if not want:
        return None, None, 0.0
    best_id = best_name = None
    best = 0.0
    for cid, cname in customers:
        have = tokens(cname)
        if not have:
            continue
        score = len(want & have) / min(len(want), len(have))
        if score > best:
            best, best_id, best_name = score, cid, cname
    return best_id, best_name, round(best, 2)


# ── Einlesen ─────────────────────────────────────────────────────────────────

def read_excel(path, sheet):
    """Alex's book, one dict per filled row."""
    ws = openpyxl.load_workbook(path, data_only=True)[sheet]
    rows = []
    for r in ws.iter_rows(min_row=5, values_only=True):
        if not r[0] or not r[2]:
            continue
        raw_name = str(r[2]).strip()
        rows.append(
            {
                "nr": int(str(r[0]).split("-")[1]),
                "raw_name": raw_name,
                "kunde": strip_prefix(raw_name),
                "typ": kind_of(raw_name),
                "leistung": r[1],
                "netto": round(float(r[3] or 0) * 100),
                "versendet": r[6],
                "faellig": r[7],
                "bezahlt": r[8],
                "offen": r[9],
                "art": (r[10] or "").strip(),
                "bemerkung": str(r[11]).strip() if r[11] else "",
            }
        )
    return rows


def read_pipe(path, arity):
    out = []
    for line in open(path, encoding="utf-8"):
        line = line.rstrip("\n")
        if not line.strip():
            continue
        f = line.split("|")
        if len(f) >= arity:
            out.append(f)
    return out


def as_date(v):
    """Alex's date cells, or None. His book holds `31.06.26` and `07..02.205`."""
    if isinstance(v, dt.datetime):
        return v.date()
    if isinstance(v, dt.date):
        return v
    return None


def raw_date_text(v):
    """The unparseable original, so it can be preserved instead of guessed at."""
    return "" if v is None or as_date(v) else str(v).strip()


# ── Abgleich ─────────────────────────────────────────────────────────────────

def compare(excel_rows, prod_rows, customers):
    """One decision per Excel row."""
    prod = {}
    for f in prod_rows:
        prod[int(f[0].split("-")[1])] = {
            "nr_text": f[0],
            "kunde": f[1],
            "typ": f[2],
            "pct": f[3],
            "status": f[4],
            "base": int(f[5]) if f[5] else None,
            "offer": int(f[6]) if f[6] else None,
            "sent": f[7] == "ja",
        }

    # Numbers prod issued that Alex's book gives to somebody else.
    decided = []
    for x in excel_rows:
        p = prod.get(x["nr"])
        cid, cname, score = best_customer(x["kunde"], customers)

        row = dict(x)
        row["match_id"] = cid
        row["match_name"] = cname or ""
        row["score"] = score
        row["prod_kunde"] = p["kunde"] if p else ""
        row["prod_status"] = p["status"] if p else ""
        row["prod_sent"] = p["sent"] if p else False

        # Which prod amount this Excel row should be compared against: a partial
        # invoice's own base, else the offer, else nothing.
        prod_netto = None
        if p:
            prod_netto = p["base"] if p["base"] is not None else p["offer"]
        row["prod_netto"] = prod_netto

        name_matches = bool(p) and score >= 0.5 and bool(tokens(x["kunde"]) & tokens(p["kunde"]))
        amount_matches = (
            prod_netto is not None and abs(prod_netto - x["netto"]) <= CENT_TOLERANCE
        )

        if not p:
            # The number is free. A customer we could not place is not a blocker —
            # creating one is the sane default, and the Kunden sheet carries that
            # single decision rather than reprinting it on every invoice row.
            row["aktion"] = "NEU"
            row["was_passiert"] = "Fehlt im System, wird als Bestandsrechnung angelegt."
            if score < 0.5:
                row["was_passiert"] += " Kunde ist neu — siehe Blatt „Kunden“."
        elif amount_matches and not name_matches:
            # Same number, same amount to the cent, different name: overwhelmingly the
            # same invoice recorded under the company where we hold the contact (or
            # the other way round), not two different invoices fighting over a number.
            row["aktion"] = "PRÜFEN"
            row["was_passiert"] = (
                f"Betrag identisch ({eur(x['netto'])}), aber im System steht "
                f"„{p['kunde']}“. Vermutlich dieselbe Rechnung — wir übernehmen "
                "deinen Namen. Falls das zwei verschiedene sind: bitte anrufen."
            )
        elif not name_matches:
            row["aktion"] = "KLÄREN" if p["sent"] else "PRÜFEN"
            row["was_passiert"] = (
                f"Nummer ist im System an „{p['kunde']}“ vergeben "
                f"({eur(p['base'] or p['offer'] or 0)}). "
                + (
                    "Diese Rechnung wurde bereits versendet — hier müssen wir "
                    "gemeinsam draufschauen."
                    if p["sent"]
                    else "Die Systemrechnung ging nie raus und wird umnummeriert; "
                         "deine Nummer gilt."
                )
            )
        elif not amount_matches:
            row["aktion"] = "PRÜFEN"
            row["was_passiert"] = (
                f"Betrag weicht ab — dein Buch {eur(x['netto'])}, "
                f"System {eur(prod_netto)}. Dein Buch gilt."
            )
        else:
            row["aktion"] = "OK"
            row["was_passiert"] = "Stimmt überein — nichts zu tun."

        if not as_date(x["bezahlt"]) and raw_date_text(x["bezahlt"]):
            row["was_passiert"] += (
                f" Bezahlt-Datum „{raw_date_text(x['bezahlt'])}“ ist unlesbar "
                "und wird als Bemerkung übernommen."
            )

        decided.append(row)

    orphan = sorted(set(prod) - {x["nr"] for x in excel_rows})
    return decided, [prod[n] for n in orphan]


# ── Schreiben ────────────────────────────────────────────────────────────────

COLUMNS = [
    ("Rg.-Nr.", 10),
    ("Aktion", 11),
    ("Kunde (dein Buch)", 30),
    ("Kunde (System)", 26),
    ("Typ", 15),
    ("Leistung", 18),
    ("Netto (dein Buch)", 16),
    ("Netto (System)", 16),
    ("Was passiert", 62),
    ("Deine Entscheidung", 20),
    ("Deine Notiz", 34),
]

DECISIONS = '"Passt so,Mein Buch gilt,System gilt,Bitte anrufen"'


def style_header(ws, ncols, row=1):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEAD
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[row].height = 28


def write_overview(wb, rows, orphans, year):
    ws = wb.create_sheet("Anleitung", 0)
    counts = {a: sum(1 for r in rows if r["aktion"] == a) for a in ACTION_FILL}
    todo = counts["PRÜFEN"] + counts["KLÄREN"]

    lines = [
        (f"Abgleich Rechnungsausgangsbuch {year}", 16, True),
        ("", 11, False),
        (f"Dein Buch hat {len(rows)} Rechnungen. Davon sind {counts['OK']} schon korrekt", 11, False),
        (f"im System, {counts['NEU']} fehlen und werden angelegt.", 11, False),
        ("", 11, False),
        (f"Bitte anschauen musst du nur {todo} Zeilen.", 12, True),
        ("", 11, False),
        ("So gehst du vor:", 12, True),
        ("1. Blatt „Abgleich“ öffnen und oben auf Spalte B filtern.", 11, False),
        ("2. Alles mit PRÜFEN oder KLÄREN durchgehen — der Rest ist erledigt.", 11, False),
        ("3. In Spalte „Deine Entscheidung“ auswählen. Vorgeschlagen ist immer", 11, False),
        ("   schon das Sinnvollste; wenn das passt, einfach „Passt so“ stehen lassen.", 11, False),
        ("4. Bei „Bitte anrufen“ kurz in die Notizspalte, worum es geht.", 11, False),
        ("5. Datei zurückschicken. Erst danach wird irgendetwas ins System geschrieben.", 11, False),
        ("", 11, False),
        ("Farben:", 12, True),
    ]
    for i, (text, size, bold) in enumerate(lines, start=1):
        c = ws.cell(row=i, column=1, value=text)
        c.font = Font(size=size, bold=bold)

    # The legend lives in one cell per line: the fill IS the swatch, so it cannot
    # drift away from its label when the column widths change.
    legend = [
        ("OK", "stimmt überein — nichts zu tun", GREEN),
        ("NEU", "fehlt im System — wird angelegt", BLUE),
        ("PRÜFEN", "Unterschied — mein Vorschlag steht in der Zeile", YELLOW),
        ("KLÄREN", "da brauchen wir dich, sonst geht es nicht", RED),
    ]
    start = len(lines) + 1
    for i, (code, meaning, fill) in enumerate(legend):
        r = start + i
        cell = ws.cell(row=r, column=1, value=f"{code}  —  {meaning}")
        cell.fill = fill
        cell.border = BORDER
        cell.font = Font(size=11)
        n = ws.cell(row=r, column=2, value=counts.get(code, 0))
        n.font = Font(bold=True, size=11)
        n.fill = fill
        n.border = BORDER
        n.alignment = Alignment(horizontal="right")

    if orphans:
        r = start + len(legend) + 2
        head = ws.cell(row=r, column=1, value="Im System, aber nicht in deinem Buch — bleibt unverändert stehen:")
        head.font = Font(bold=True, size=11)
        for i, o in enumerate(orphans, start=1):
            ws.cell(row=r + i, column=1, value=f"{o['nr_text']}   {o['kunde']}")

    for col, width in [("A", 72), ("B", 8)]:
        ws.column_dimensions[col].width = width
    return ws


def write_compare(wb, rows):
    ws = wb.create_sheet("Abgleich")
    for i, (title, width) in enumerate(COLUMNS, start=1):
        ws.cell(row=1, column=i, value=title)
        ws.column_dimensions[get_column_letter(i)].width = width
    style_header(ws, len(COLUMNS))

    for i, r in enumerate(rows, start=2):
        vals = [
            f"2026-{r['nr']:02d}",
            r["aktion"],
            r["raw_name"],
            r["prod_kunde"] or (r["match_name"] if r["score"] >= 0.5 else ""),
            {"full": "Rechnung", "partial_first": "Anzahlung",
             "partial_final": "Restzahlung", "gutschrift": "Gutschrift",
             "lagerung": "Lagerung"}[r["typ"]],
            r["leistung"] if isinstance(r["leistung"], str) else (
                as_date(r["leistung"]).strftime("%d.%m.%Y") if as_date(r["leistung"]) else ""
            ),
            r["netto"] / 100,
            r["prod_netto"] / 100 if r["prod_netto"] is not None else None,
            r["was_passiert"],
            "Passt so",
            "",
        ]
        for c, v in enumerate(vals, start=1):
            cell = ws.cell(row=i, column=c, value=v)
            cell.border = BORDER
            cell.fill = ACTION_FILL[r["aktion"]]
            if c in (7, 8):
                cell.number_format = '#,##0.00\\ "€"'
            if c == 9:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            if c == 2:
                cell.font = Font(bold=True, size=10)
        ws.row_dimensions[i].height = 30

    last = len(rows) + 1
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{last}"
    ws.freeze_panes = "A2"

    dv = DataValidation(type="list", formula1=DECISIONS, allow_blank=False)
    dv.error = "Bitte einen der vier Punkte auswählen."
    dv.errorTitle = "Ungültige Auswahl"
    ws.add_data_validation(dv)
    dv.add(f"J2:J{last}")

    # A decision other than the suggested one stands out, so nothing he changed
    # gets lost among 86 rows on the way back to us.
    ws.conditional_formatting.add(
        f"A2:K{last}",
        FormulaRule(formula=['$J2<>"Passt so"'],
                    fill=PatternFill("solid", fgColor="FFE0B2")),
    )
    return ws


def write_customers(wb, rows):
    """Only the names that could not be matched — one decision each, not 86."""
    unclear = {}
    for r in rows:
        if r["score"] < 0.5:
            unclear.setdefault(r["kunde"], []).append(r["nr"])

    ws = wb.create_sheet("Kunden")
    headers = [("Kunde (dein Buch)", 36), ("Betrifft Rechnungen", 26),
               ("Vorschlag System", 30), ("Deine Entscheidung", 24), ("Deine Notiz", 34)]
    for i, (title, width) in enumerate(headers, start=1):
        ws.cell(row=1, column=i, value=title)
        ws.column_dimensions[get_column_letter(i)].width = width
    style_header(ws, len(headers))

    for i, (name, nrs) in enumerate(sorted(unclear.items()), start=2):
        ws.cell(row=i, column=1, value=name)
        ws.cell(row=i, column=2, value=", ".join(f"2026-{n:02d}" for n in sorted(nrs)))
        ws.cell(row=i, column=3, value="— kein passender gefunden —")
        ws.cell(row=i, column=4, value="Neu anlegen")
        for c in range(1, 6):
            ws.cell(row=i, column=c).border = BORDER
            ws.cell(row=i, column=c).fill = RED

    if unclear:
        dv = DataValidation(
            type="list", formula1='"Neu anlegen,Ist ein bestehender Kunde,Bitte anrufen"',
            allow_blank=False,
        )
        ws.add_data_validation(dv)
        dv.add(f"D2:D{len(unclear) + 1}")
        ws.freeze_panes = "A2"
    return ws


def main():
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--excel", required=True)
    ap.add_argument("--sheet", default="2026")
    ap.add_argument("--prod-invoices", required=True)
    ap.add_argument("--prod-customers", required=True)
    ap.add_argument("--out", required=True)
    args = ap.parse_args()

    excel_rows = read_excel(args.excel, args.sheet)
    prod_rows = read_pipe(args.prod_invoices, 8)
    customers = [(f[0], f[1]) for f in read_pipe(args.prod_customers, 2)]

    rows, orphans = compare(excel_rows, prod_rows, customers)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    write_overview(wb, rows, orphans, args.sheet)
    write_compare(wb, rows)
    write_customers(wb, rows)
    wb.save(args.out)

    counts = {a: sum(1 for r in rows if r["aktion"] == a) for a in ACTION_FILL}
    print(f"{args.out}")
    print(f"  {len(rows)} Zeilen aus dem Buch, {len(prod_rows)} im System")
    for a in ("OK", "NEU", "PRÜFEN", "KLÄREN"):
        print(f"  {a:8} {counts[a]}")
    print(f"  Alex muss anschauen: {counts['PRÜFEN'] + counts['KLÄREN']}")


if __name__ == "__main__":
    main()
